from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from .models import Contact, Tag, Campaign
from . import db

bp = Blueprint('main', __name__)

@bp.route('/')
def index():
    return redirect(url_for('main.dashboard'))

@bp.route('/dashboard')
def dashboard():
    # Calculate stats
    total_contacts = Contact.query.count()

    today = datetime.utcnow().date()
    first_day_of_month = today.replace(day=1)
    new_contacts_monthly = Contact.query.filter(Contact.created_at >= first_day_of_month).count()

    campaigns_sent = Campaign.query.filter(Campaign.recipients.any()).count()

    stats = {
        'total_contacts': total_contacts,
        'new_contacts_monthly': new_contacts_monthly,
        'campaigns_sent': campaigns_sent
    }

    # Fetch recent campaigns
    recent_campaigns = Campaign.query.filter(Campaign.recipients.any()).order_by(Campaign.created_at.desc()).limit(5).all()

    return render_template('dashboard.html', stats=stats, recent_campaigns=recent_campaigns)

@bp.route('/contacts')
def list_contacts():
    contacts = Contact.query.order_by(Contact.created_at.desc()).all()
    return render_template('contacts.html', contacts=contacts)

@bp.route('/campaigns/new', methods=('GET', 'POST'))
def new_campaign():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'save_draft':
            name = request.form.get('campaign_name')
            message = request.form.get('message')
            tags_string = request.form.get('tags', '')

            if not name or not message:
                flash('اسم الحملة ونص الرسالة حقول إلزامية.', 'danger')
                return render_template('new_campaign.html')

            # Create the campaign
            new_campaign = Campaign(name=name, message=message)
            db.session.add(new_campaign) # Add to session before creating relationships

            # Handle tags
            if tags_string:
                tag_names = [name.strip() for name in tags_string.split(',') if name.strip()]
                for tag_name in tag_names:
                    tag = Tag.query.filter_by(name=tag_name).first()
                    if not tag:
                        tag = Tag(name=tag_name)
                        db.session.add(tag)
                    new_campaign.tags.append(tag)

            db.session.commit()

            flash(f'تم حفظ الحملة "{name}" كمسودة بنجاح!', 'success')
            return redirect(url_for('main.new_campaign'))

        # The 'export' action will be handled in the next step
        elif action == 'export':
            name = request.form.get('campaign_name')
            message = request.form.get('message')
            tags_string = request.form.get('tags', '')
            use_shield = request.form.get('anti_spam_shield_active') == 'on'

            if not name or not message:
                flash('اسم الحملة ونص الرسالة حقول إلزامية للتصدير.', 'danger')
                return render_template('new_campaign.html')

            # Find or create campaign object
            campaign = Campaign.query.filter_by(name=name).first()
            if not campaign:
                campaign = Campaign(name=name, message=message)
                db.session.add(campaign)
            else: # if campaign exists, update its message
                campaign.message = message

            # Find or create tag objects and associate them with the campaign
            campaign.tags.clear()
            tag_names = [name.strip() for name in tags_string.split(',') if name.strip()]
            for tag_name in tag_names:
                tag = Tag.query.filter_by(name=tag_name).first()
                if not tag:
                    tag = Tag(name=tag_name)
                    db.session.add(tag)
                campaign.tags.append(tag)

            # Determine final recipient list
            query = Contact.query
            if tag_names:
                for tag_name in tag_names:
                    query = query.filter(Contact.tags.any(name=tag_name))

            if use_shield:
                # Get IDs of contacts who have already received this campaign
                existing_recipient_ids = [r.id for r in campaign.recipients]
                if existing_recipient_ids:
                    query = query.filter(Contact.id.notin_(existing_recipient_ids))

            final_recipients = query.all()

            if not final_recipients:
                flash('لا يوجد مستلمين جدد لهذه الحملة.', 'warning')
                return redirect(url_for('main.new_campaign'))

            # Log new recipients and prepare data for Excel
            excel_data = []
            for contact in final_recipients:
                campaign.recipients.append(contact)
                personalized_message = message.replace('[الاسم]', contact.name)
                excel_data.append([contact.phone, personalized_message])

            db.session.commit()

            # Create Excel file in memory
            wb = Workbook()
            ws = wb.active
            ws.title = "Campaign Export"
            ws.append(['PhoneNumber', 'Message']) # Add headers
            for row in excel_data:
                ws.append(row)

            # Save to a virtual file
            target = BytesIO()
            wb.save(target)
            target.seek(0)

            flash(f'تم تصدير {len(final_recipients)} جهة اتصال بنجاح!', 'success')

            return send_file(
                target,
                as_attachment=True,
                download_name=f'{name}_export.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    return render_template('new_campaign.html')

@bp.route('/contacts/<int:contact_id>/edit', methods=('GET', 'POST'))
def edit_contact(contact_id):
    contact = db.get_or_404(Contact, contact_id)

    if request.method == 'POST':
        # Update contact details
        contact.name = request.form['name']
        new_phone = request.form['phone']
        contact.source = request.form['source']

        # Check for phone number uniqueness
        existing_contact_with_phone = Contact.query.filter(Contact.phone == new_phone, Contact.id != contact_id).first()
        if existing_contact_with_phone:
            flash(f'رقم الهاتف {new_phone} مسجل بالفعل لجهة اتصال أخرى.', 'danger')
        else:
            contact.phone = new_phone

            # Handle tag management
            tags_string = request.form.get('tags', '').strip()
            if tags_string:
                tag_names = [name.strip() for name in tags_string.split(',') if name.strip()]
                for name in tag_names:
                    tag = Tag.query.filter_by(name=name).first()
                    if not tag:
                        tag = Tag(name=name)
                        db.session.add(tag)
                    if tag not in contact.tags:
                        contact.tags.append(tag)

            db.session.commit()
            flash('تم تحديث بيانات جهة الاتصال بنجاح!', 'success')

        return redirect(url_for('main.edit_contact', contact_id=contact.id))

    return render_template('edit_contact.html', contact=contact)

# Tag Management Routes
@bp.route('/tags', methods=('GET', 'POST'))
def tags_index():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if name:
            existing_tag = Tag.query.filter_by(name=name).first()
            if not existing_tag:
                new_tag = Tag(name=name)
                db.session.add(new_tag)
                db.session.commit()
                flash(f'تم إنشاء الوسم "{name}" بنجاح.', 'success')
            else:
                flash(f'الوسم "{name}" موجود بالفعل.', 'warning')
        return redirect(url_for('main.tags_index'))

    tags = Tag.query.order_by(Tag.name).all()
    return render_template('tags/index.html', tags=tags)

@bp.route('/tags/<int:tag_id>/edit', methods=('GET', 'POST'))
def edit_tag(tag_id):
    tag = db.get_or_404(Tag, tag_id)
    if request.method == 'POST':
        new_name = request.form.get('name', '').strip()
        if new_name:
            existing_tag = Tag.query.filter(Tag.name == new_name, Tag.id != tag_id).first()
            if not existing_tag:
                tag.name = new_name
                db.session.commit()
                flash('تم تحديث اسم الوسم بنجاح.', 'success')
                return redirect(url_for('main.tags_index'))
            else:
                flash(f'الوسم "{new_name}" موجود بالفعل.', 'danger')
        else:
            flash('اسم الوسم لا يمكن أن يكون فارغًا.', 'danger')
        return redirect(url_for('main.edit_tag', tag_id=tag.id))

    return render_template('tags/edit.html', tag=tag)

@bp.route('/tags/<int:tag_id>/delete', methods=['POST'])
def delete_tag(tag_id):
    tag = db.get_or_404(Tag, tag_id)
    # Dissociate from all contacts before deleting
    tag.contacts = []
    db.session.delete(tag)
    db.session.commit()
    flash(f'تم حذف الوسم "{tag.name}" بنجاح.', 'success')
    return redirect(url_for('main.tags_index'))


@bp.route('/contacts/import', methods=['GET', 'POST'])
def import_contacts():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('لم يتم العثور على جزء الملف.', 'danger')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('لم يتم تحديد أي ملف.', 'danger')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            try:
                tags_string = request.form.get('tags', '')
                tag_objects = []
                if tags_string:
                    tag_names = [name.strip() for name in tags_string.split(',') if name.strip()]
                    for tag_name in tag_names:
                        tag = Tag.query.filter_by(name=tag_name).first()
                        if not tag:
                            tag = Tag(name=tag_name)
                            db.session.add(tag)
                        tag_objects.append(tag)

                try:
                    from openpyxl import load_workbook
                    workbook = load_workbook(filename=file)
                except Exception as e:
                    flash(f'خطأ في قراءة ملف Excel: {e}', 'danger')
                    return redirect(request.url)

                sheet = workbook.active

                imported_count = 0
                skipped_count = 0

                existing_phones = {c.phone for c in Contact.query.all()}

                for row in sheet.iter_rows(min_row=2, values_only=True): # Skip header row
                    name, phone, source = row[0], row[1], row[2]

                    if not phone or str(phone) in existing_phones:
                        skipped_count += 1
                        continue

                    new_contact = Contact(name=str(name), phone=str(phone), source=str(source or ''))
                    new_contact.tags.extend(tag_objects)
                    db.session.add(new_contact)

                    existing_phones.add(str(phone))
                    imported_count += 1

                db.session.commit()
                flash(f'تم استيراد {imported_count} جهة اتصال جديدة بنجاح. تم تخطي {skipped_count} صف (مكرر أو بدون رقم هاتف).', 'success')

            except Exception as e:
                db.session.rollback()
                flash(f'حدث خطأ غير متوقع: {e}', 'danger')

            return redirect(url_for('main.list_contacts'))

        else:
            flash('صيغة الملف غير مدعومة. الرجاء تحميل ملف .xlsx فقط.', 'danger')
            return redirect(request.url)

    return render_template('contacts/import.html')


@bp.route('/contacts/new', methods=('GET', 'POST'))
def add_contact():
    # This page is now secondary, but the logic is still needed.
    # It might be converted to a modal dialog in the future.
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        source = request.form.get('source') # .get() is safer for optional fields

        error = None

        if not name:
            error = 'الاسم مطلوب.'
        elif not phone:
            error = 'رقم الهاتف مطلوب.'

        # Check if phone number already exists
        if error is None:
            existing_contact = Contact.query.filter_by(phone=phone).first()
            if existing_contact:
                error = f'رقم الهاتف {phone} مسجل بالفعل.'

        if error is None:
            # No errors, so create the contact and save to DB
            new_contact = Contact(name=name, phone=phone, source=source)
            db.session.add(new_contact)
            db.session.commit()
            flash('تمت إضافة جهة الاتصال بنجاح!', 'success')
            return redirect(url_for('main.add_contact'))

        # If there was an error, show it to the user
        flash(error, 'danger')

    return render_template('add_contact.html')
